import os
import tabula
import pandas as pd
from tkinter import Tk, filedialog
from openpyxl import load_workbook

# Function to select a folder using a dialog
def select_folder():
    folder_path = filedialog.askdirectory(title="Select Folder Containing PDFs")
    return folder_path

# Function to select the template JSON file using a dialog
def select_template():
    template_path = filedialog.askopenfilename(title="Select Tabula Template", filetypes=(("JSON Files", "*.json"),))
    return template_path

# Function to select the output Excel file path and name using a dialog
def select_output_excel():
    output_file = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=(("Excel Files", "*.xlsx"),), title="Save Output as Excel")
    return output_file

# Initialize the Tkinter root window (this will not be shown)
root = Tk()
root.withdraw()  # Hide the root window

# Select the folder containing PDF files
folder_path = select_folder()
if not folder_path:
    print("No folder selected, exiting.")
    exit()

# Select the template JSON file
template_path = select_template()
if not template_path:
    print("No template selected, exiting.")
    exit()

# Select the output Excel file
output_file = select_output_excel()
if not output_file:
    print("No output file selected, exiting.")
    exit()

# Initialize the Excel writer
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    # Start writing to the Excel file
    row_index = 0  # Start at the first row

    # Loop through each PDF file in the selected folder
    for idx, filename in enumerate(os.listdir(folder_path)):
        if filename.endswith('.pdf'):
            input_path = os.path.join(folder_path, filename)

            try:
                # Extract tables from the current PDF using the template
                print(f"Processing {filename}...")
                tables = tabula.read_pdf_with_template(input_path=input_path, template_path=template_path)

                # If tables are found, process them and store them in columns
                if tables:
                    print(f"Extracted tables from {filename}:")
                    
                    # Access the workbook and worksheet
                    workbook = writer.book
                    if 'All_Tables' not in writer.sheets:
                        worksheet = workbook.create_sheet('All_Tables')
                        writer.sheets['All_Tables'] = worksheet
                    else:
                        worksheet = writer.sheets['All_Tables']

                    # Write the file name in the first column (column A)
                    worksheet.cell(row=row_index + 1, column=1, value=filename)

                    # Loop through the tables extracted from the current PDF
                    for table_idx, table in enumerate(tables, start=1):
                        print(f"Table {table_idx} from {filename}:")
                        print(table)  # Print each extracted table
                        
                        # Write the table to the Excel sheet at the appropriate position (row, column)
                        table.to_excel(writer, sheet_name='All_Tables', startrow=row_index, startcol=table_idx, index=False, header=True)

                    # Add a hyperlink to the PDF file in the last column
                    last_col = len(tables) + 2  # Adjust the column index for the hyperlink
                    worksheet.cell(row=row_index + 1, column=last_col).value = f'=HYPERLINK("{input_path}", "Open PDF")'

                    # After storing all tables for this PDF, move to the next row
                    row_index += len(tables[0]) + 1  # Increment row index based on the first table's row count
                else:
                    print(f"No tables found in {filename}")
            
            except Exception as e:
                print(f"An error occurred while processing {filename}: {e}")

print(f"All extracted data has been saved to {output_file}")
